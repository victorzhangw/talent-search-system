"""
Shared logic for importing trait data from Excel.
Used by both the one-off migration script and the /api/admin/traits/upload endpoint.
"""

import io
import json
import os
import re
from datetime import datetime

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from sqlalchemy import text
from ..database.connection import get_db_engine

BAND_SHEET_KEY = 'TraitSemanticBands'
INTERACTION_SHEET_KEY = 'interaction_narrative'

BACKUP_DIR = os.path.join(os.path.dirname(__file__), '..', '..', '..', 'scripts', 'backups')

COL = {
    'trait_id':                0,
    'trait_name_zh':           1,
    'trait_name_en':           2,
    'dimension_group':         3,
    'definition_zh':           4,
    'definition_en':           5,
    'hidden_anchor':           6,
    'band':                    7,
    'band_range':              8,
    'semantic_label':          9,
    'semantic_description':    10,
    'management_focus':        11,
    'usage_note':              12,
    'trait_interaction_guide': 13,
    'report_wording_zh':       14,
    'report_wording_friendly': 15,
    'ai_do':                   16,
    'ai_dont':                 17,
    'version':                 18,
}


def _cell(row, idx):
    try:
        v = row[idx]
        if v is None:
            return None
        return str(v).strip() if not isinstance(v, str) else v.strip()
    except IndexError:
        return None


def _parse_band_range(s):
    if not s:
        return None, None
    m = re.match(r'(\d+)\s*[-~]\s*(\d+)', str(s).strip())
    return (int(m.group(1)), int(m.group(2))) if m else (None, None)


def _extract_project(trait_id):
    return trait_id.split('_')[0] if trait_id and '_' in trait_id else None


def _is_valid_trait_id(trait_id):
    return bool(trait_id and re.match(r'^[A-Z]{2,6}_\d{2,3}$', str(trait_id).strip()))


def _row_has_data(row):
    return any(c is not None and str(c).strip() != '' for c in row)


def _is_template_row(row):
    """Template/instruction rows (e.g. a '// 系統顯示名稱' guide row under the header) aren't real data."""
    return any(isinstance(c, str) and c.strip().startswith('//') for c in row)


def _parse_band_sheet(ws):
    definitions = {}
    bands = []
    skipped = []
    rows = list(ws.iter_rows(values_only=True))
    total_rows = 0
    for excel_row, row in enumerate(rows[1:], start=2):
        if not _row_has_data(row) or _is_template_row(row):
            continue
        total_rows += 1
        trait_id = _cell(row, COL['trait_id'])
        band = _cell(row, COL['band'])
        if not _is_valid_trait_id(trait_id):
            skipped.append({'row': excel_row, 'reason': f'無效的 trait_id: {trait_id!r}'})
            continue
        if band not in ('A', 'B', 'C'):
            skipped.append({'row': excel_row, 'trait_id': trait_id, 'reason': f'無效的 band 值（需為 A/B/C）: {band!r}'})
            continue
        if trait_id not in definitions:
            definitions[trait_id] = {
                'trait_id':      trait_id,
                'name_zh':       _cell(row, COL['trait_name_zh']),
                'name_en':       _cell(row, COL['trait_name_en']),
                'dimension':     _cell(row, COL['dimension_group']),
                'definition':    _cell(row, COL['definition_zh']),
                'definition_en': _cell(row, COL['definition_en']),
                'hidden_anchor': _cell(row, COL['hidden_anchor']),
            }
        min_score, max_score = _parse_band_range(_cell(row, COL['band_range']))
        ai_do = [s.strip() for s in re.split(r'[;\n]+', _cell(row, COL['ai_do']) or '') if s.strip()]
        ai_dont = [s.strip() for s in re.split(r'[;\n]+', _cell(row, COL['ai_dont']) or '') if s.strip()]
        bands.append({
            'trait_id':               trait_id,
            'band':                   band,
            'min_score':              min_score,
            'max_score':              max_score,
            'semantic_label':         _cell(row, COL['semantic_label']),
            'description':            _cell(row, COL['semantic_description']),
            'management_focus':       _cell(row, COL['management_focus']),
            'usage_note':             _cell(row, COL['usage_note']),
            'trait_interaction_guide': _cell(row, COL['trait_interaction_guide']),
            'report_wording':         _cell(row, COL['report_wording_zh']),
            'report_wording_friendly': _cell(row, COL['report_wording_friendly']),
            'ai_guidance':            {'do': ai_do, 'dont': ai_dont},
            'version':                _cell(row, COL['version']),
            'trait_project':          _extract_project(trait_id),
        })
    return list(definitions.values()), bands, total_rows, skipped


def _parse_interaction_sheet(ws):
    interactions = []
    skipped = []
    rows = list(ws.iter_rows(values_only=True))
    total_rows = 0
    for excel_row, row in enumerate(rows[1:], start=2):
        if not _row_has_data(row) or _is_template_row(row):
            continue
        total_rows += 1
        primary_trait_id = _cell(row, 0)
        if not primary_trait_id:
            skipped.append({'row': excel_row, 'reason': '缺少 primary_trait_id'})
            continue
        trigger_trait_id = trigger_band = None
        try:
            data = json.loads(_cell(row, 4) or '{}')
            t = data.get('trigger', [])
            if t:
                trigger_trait_id = t[0].get('id')
                trigger_band = t[0].get('band')
        except (json.JSONDecodeError, AttributeError):
            pass
        interactions.append({
            'primary_trait_id': primary_trait_id,
            'primary_band':     _cell(row, 2),
            'trigger_trait_id': trigger_trait_id,
            'trigger_band':     trigger_band,
            'narrative':        _cell(row, 5),
        })
    return interactions, total_rows, skipped


def _find_sheet(wb, key):
    for name in wb.sheetnames:
        if key.lower() in name.lower():
            return name
    return None


def parse_excel_bytes(file_bytes):
    """
    Parse Excel from bytes (for API upload).
    Returns (definitions, bands, interactions, error, mismatch).
    `mismatch` is None when parsed row counts match the Excel data row counts,
    otherwise a dict describing which rows were skipped and why.
    """
    if not OPENPYXL_AVAILABLE:
        return None, None, None, 'openpyxl not installed on server', None
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        return None, None, None, f'Cannot open Excel: {e}', None

    band_name = _find_sheet(wb, BAND_SHEET_KEY)
    interaction_name = _find_sheet(wb, INTERACTION_SHEET_KEY)
    missing = [k for k, v in {BAND_SHEET_KEY: band_name, INTERACTION_SHEET_KEY: interaction_name}.items() if not v]
    if missing:
        return None, None, None, f'Cannot find sheets for: {missing}. Found: {wb.sheetnames}', None

    try:
        definitions, bands, band_total, band_skipped = _parse_band_sheet(wb[band_name])
        interactions, interaction_total, interaction_skipped = _parse_interaction_sheet(wb[interaction_name])
    except Exception as e:
        return None, None, None, f'Parse error: {e}', None
    finally:
        wb.close()

    mismatch = None
    if band_skipped or interaction_skipped:
        parts = []
        if band_skipped:
            parts.append(
                f'TraitSemanticBands：Excel 有 {band_total} 筆資料，成功解析 {len(bands)} 筆，'
                f'跳過 {len(band_skipped)} 筆（' +
                '；'.join(f"第{s['row']}列：{s['reason']}" for s in band_skipped[:10]) +
                ('…' if len(band_skipped) > 10 else '') + '）'
            )
        if interaction_skipped:
            parts.append(
                f'interaction_narrative：Excel 有 {interaction_total} 筆資料，成功解析 {len(interactions)} 筆，'
                f'跳過 {len(interaction_skipped)} 筆（' +
                '；'.join(f"第{s['row']}列：{s['reason']}" for s in interaction_skipped[:10]) +
                ('…' if len(interaction_skipped) > 10 else '') + '）'
            )
        mismatch = {
            'message': '匯入筆數與 Excel 資料筆數不符，已拒絕匯入。' + ' '.join(parts),
            'band_total': band_total,
            'band_parsed': len(bands),
            'band_skipped': band_skipped,
            'interaction_total': interaction_total,
            'interaction_parsed': len(interactions),
            'interaction_skipped': interaction_skipped,
        }

    return definitions, bands, interactions, None, mismatch


def _escape_sql_value(v):
    if v is None:
        return 'NULL'
    if isinstance(v, (int, float)):
        return str(v)
    if isinstance(v, dict):
        s = json.dumps(v, ensure_ascii=False).replace("'", "''")
        return f"'{s}'"
    return "'" + str(v).replace("'", "''") + "'"


def create_backup_sql(backup_dir=None):
    """Dump the three trait tables to a timestamped SQL file. Returns file path."""
    target_dir = os.path.abspath(backup_dir or BACKUP_DIR)
    os.makedirs(target_dir, exist_ok=True)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    path = os.path.join(target_dir, f'trait_backup_{ts}.sql')

    engine = get_db_engine()
    with engine.connect() as conn:
        with open(path, 'w', encoding='utf-8') as f:
            f.write(f"-- Trait tables backup {ts}\nBEGIN;\n")
            for table in ('trait_definitions', 'trait_bands', 'trait_interactions'):
                result = conn.execute(text(f"SELECT * FROM {table}"))
                cols = list(result.keys())
                rows = result.fetchall()
                f.write(f"\n-- {table} ({len(rows)} rows)\nDELETE FROM {table};\n")
                for row in rows:
                    vals = ', '.join(_escape_sql_value(row[c]) for c in cols)
                    f.write(f"INSERT INTO {table} ({', '.join(cols)}) VALUES ({vals});\n")
            f.write("\nCOMMIT;\n")
    return path


def apply_schema_migration():
    """Add new columns if not yet present (idempotent)."""
    sqls = [
        "ALTER TABLE trait_definitions ADD COLUMN IF NOT EXISTS definition_en TEXT",
        "ALTER TABLE trait_definitions ADD COLUMN IF NOT EXISTS hidden_anchor TEXT",
        "ALTER TABLE trait_bands ADD COLUMN IF NOT EXISTS usage_note TEXT",
        "ALTER TABLE trait_bands ADD COLUMN IF NOT EXISTS trait_interaction_guide TEXT",
        "ALTER TABLE trait_bands ADD COLUMN IF NOT EXISTS version VARCHAR(20)",
    ]
    engine = get_db_engine()
    with engine.connect() as conn:
        for sql in sqls:
            conn.execute(text(sql))
        conn.commit()


def write_traits_to_db(definitions, bands, interactions):
    """Truncate the three tables and bulk-insert parsed data."""
    engine = get_db_engine()
    with engine.begin() as conn:
        conn.execute(text("TRUNCATE trait_interactions, trait_bands, trait_definitions CASCADE"))

        for d in definitions:
            conn.execute(text("""
                INSERT INTO trait_definitions
                    (trait_id, name_zh, name_en, dimension, definition, definition_en, hidden_anchor)
                VALUES (:trait_id, :name_zh, :name_en, :dimension, :definition, :definition_en, :hidden_anchor)
            """), d)

        for b in bands:
            b2 = dict(b)
            b2['ai_guidance'] = json.dumps(b2['ai_guidance'], ensure_ascii=False)
            conn.execute(text("""
                INSERT INTO trait_bands
                    (trait_id, band, min_score, max_score, semantic_label, description,
                     management_focus, usage_note, trait_interaction_guide,
                     report_wording, report_wording_friendly, trait_project, ai_guidance, version)
                VALUES
                    (:trait_id, :band, :min_score, :max_score, :semantic_label, :description,
                     :management_focus, :usage_note, :trait_interaction_guide,
                     :report_wording, :report_wording_friendly, :trait_project,
                     CAST(:ai_guidance AS jsonb), :version)
            """), b2)

        for i in interactions:
            conn.execute(text("""
                INSERT INTO trait_interactions
                    (primary_trait_id, primary_band, trigger_trait_id, trigger_band, narrative)
                VALUES (:primary_trait_id, :primary_band, :trigger_trait_id, :trigger_band, :narrative)
            """), i)


def list_backups(backup_dir=None):
    """Return list of backup file metadata sorted newest-first."""
    target_dir = os.path.abspath(backup_dir or BACKUP_DIR)
    if not os.path.isdir(target_dir):
        return []
    files = []
    for name in os.listdir(target_dir):
        if name.startswith('trait_backup_') and name.endswith('.sql'):
            full = os.path.join(target_dir, name)
            files.append({
                'filename': name,
                'size_bytes': os.path.getsize(full),
                'created_at': datetime.fromtimestamp(os.path.getmtime(full)).isoformat(),
            })
    return sorted(files, key=lambda x: x['created_at'], reverse=True)


def restore_from_backup(filename, backup_dir=None):
    """Execute a backup SQL file against the DB."""
    target_dir = os.path.abspath(backup_dir or BACKUP_DIR)
    path = os.path.join(target_dir, filename)
    if not os.path.isfile(path):
        raise FileNotFoundError(f'Backup not found: {filename}')
    # Sanitise filename (no path traversal)
    if '..' in filename or '/' in filename or '\\' in filename:
        raise ValueError('Invalid filename')

    with open(path, 'r', encoding='utf-8') as f:
        sql = f.read()

    engine = get_db_engine()
    with engine.begin() as conn:
        # Execute statement by statement, skip empty lines and comments
        statements = [s.strip() for s in sql.split('\n') if s.strip() and not s.strip().startswith('--')]
        batch = []
        for stmt in statements:
            batch.append(stmt)
            if stmt.rstrip().endswith(';'):
                full = ' '.join(batch)
                if full.upper() not in ('BEGIN;', 'COMMIT;'):
                    conn.execute(text(full))
                batch = []
