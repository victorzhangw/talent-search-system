"""
Migrate trait definitions from Excel spec to PostgreSQL.

Usage:
    python scripts/migrate_traits_from_excel.py --excel path/to/file.xlsx [--dry-run]

Steps:
    1. Parse Excel (sheets: 02_TraitSemanticBands, 08 interaction_narrative)
    2. Backup existing DB data to SQL file under scripts/backups/
    3. TRUNCATE and re-insert all three trait tables
"""

import sys
import os
import argparse
import json
import re
from datetime import datetime

# Allow importing from api_v2 package
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

try:
    import openpyxl
except ImportError:
    print("[ERROR] openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# Load .env before importing DB modules
from dotenv import load_dotenv
env_path = os.path.join(os.path.dirname(__file__), '..', 'api_v2', '.env')
load_dotenv(env_path, encoding='utf-8-sig')

from api_v2.database.connection import get_db_engine


# ---------------------------------------------------------------------------
# Excel parsing
# ---------------------------------------------------------------------------

BAND_SHEET_KEY = 'TraitSemanticBands'
INTERACTION_SHEET_KEY = 'interaction_narrative'
ENDPOINT_SHEET_KEY = '09_endpoints'
BLOCK_SHEET_KEY = '10_endpoint_blocks'

# DDL for the endpoint tables lives in one place; this script executes it so a
# fresh environment can be set up from the script alone.
ENDPOINT_DDL_PATH = os.path.join(
    os.path.dirname(__file__), 'migrations', '2026-08-05_add_endpoint_tables.sql')

# Expected column indices (0-based) for Sheet 2
COL = {
    'trait_id':                0,
    'trait_name_zh':           1,
    'trait_name_en':           2,
    'dimension_group':         3,
    'definition_zh':           4,
    'definition_en':           5,
    'hidden_anchor':           6,
    'band':                    7,
    'band_range':              8,
    'semantic_label':          9,
    'semantic_description':    10,
    'management_focus':        11,
    'usage_note':              12,
    'trait_interaction_guide': 13,
    'report_wording_zh':       14,
    'report_wording_friendly': 15,
    'ai_do':                   16,
    'ai_dont':                 17,
    'version':                 18,
}


def _cell(row, idx):
    """Safe cell value access."""
    try:
        v = row[idx]
        if v is None:
            return None
        return str(v).strip() if not isinstance(v, str) else v.strip()
    except IndexError:
        return None


def _parse_band_range(band_range_str):
    """Parse '76-100' → (76, 100). Returns (None, None) on failure."""
    if not band_range_str:
        return None, None
    m = re.match(r'(\d+)\s*[-~]\s*(\d+)', str(band_range_str).strip())
    if m:
        return int(m.group(1)), int(m.group(2))
    return None, None


def _extract_project(trait_id):
    """'CIA_01' → 'CIA'"""
    if trait_id and '_' in trait_id:
        return trait_id.split('_')[0]
    return None


def _is_valid_trait_id(trait_id):
    """Accept only IDs like CIA_01, ANI_10 — skip header/comment rows."""
    return bool(trait_id and re.match(r'^[A-Z]{2,6}_\d{2,3}$', str(trait_id).strip()))


def parse_band_sheet(ws):
    """
    Returns:
        definitions: dict[trait_id] → dict  (one per unique trait)
        bands:       list of dict            (one per trait × band)
    """
    definitions = {}
    bands = []

    rows = list(ws.iter_rows(values_only=True))
    header = rows[0] if rows else []

    for row in rows[1:]:
        trait_id = _cell(row, COL['trait_id'])
        if not _is_valid_trait_id(trait_id):
            continue

        band = _cell(row, COL['band'])
        if not band or band not in ('A', 'B', 'C'):
            continue

        # Accumulate one definition per trait_id (first occurrence wins)
        if trait_id not in definitions:
            definitions[trait_id] = {
                'trait_id':     trait_id,
                'name_zh':      _cell(row, COL['trait_name_zh']),
                'name_en':      _cell(row, COL['trait_name_en']),
                'dimension':    _cell(row, COL['dimension_group']),
                'definition':   _cell(row, COL['definition_zh']),
                'definition_en': _cell(row, COL['definition_en']),
                'hidden_anchor': _cell(row, COL['hidden_anchor']),
            }

        min_score, max_score = _parse_band_range(_cell(row, COL['band_range']))

        ai_do_raw = _cell(row, COL['ai_do']) or ''
        ai_dont_raw = _cell(row, COL['ai_dont']) or ''
        ai_do_list = [s.strip() for s in re.split(r'[;\n]+', ai_do_raw) if s.strip()]
        ai_dont_list = [s.strip() for s in re.split(r'[;\n]+', ai_dont_raw) if s.strip()]

        bands.append({
            'trait_id':               trait_id,
            'band':                   band,
            'min_score':              min_score,
            'max_score':              max_score,
            'semantic_label':         _cell(row, COL['semantic_label']),
            'description':            _cell(row, COL['semantic_description']),
            'management_focus':       _cell(row, COL['management_focus']),
            'usage_note':             _cell(row, COL['usage_note']),
            'trait_interaction_guide': _cell(row, COL['trait_interaction_guide']),
            'report_wording':         _cell(row, COL['report_wording_zh']),
            'report_wording_friendly': _cell(row, COL['report_wording_friendly']),
            'ai_guidance':            {'do': ai_do_list, 'dont': ai_dont_list},
            'version':                _cell(row, COL['version']),
            'trait_project':          _extract_project(trait_id),
        })

    return list(definitions.values()), bands


def parse_interaction_sheet(ws):
    """Returns list of interaction dicts."""
    interactions = []
    rows = list(ws.iter_rows(values_only=True))

    for row in rows[1:]:
        primary_trait_id = _cell(row, 0)
        if not primary_trait_id:
            continue
        raw_primary_band = _cell(row, 2)
        # Strip Chinese descriptor suffix, e.g. 'A (高)' -> 'A', so it matches
        # trait_bands.band and trigger_band's plain A/B/C format.
        primary_band = raw_primary_band.split('(')[0].strip() if raw_primary_band else raw_primary_band
        interaction_json_str = _cell(row, 4)
        narrative = _cell(row, 5)

        trigger_trait_id = None
        trigger_band = None
        if interaction_json_str:
            try:
                data = json.loads(interaction_json_str)
                triggers = data.get('trigger', [])
                if triggers:
                    trigger_trait_id = triggers[0].get('id')
                    trigger_band = triggers[0].get('band')
            except (json.JSONDecodeError, AttributeError):
                pass

        interactions.append({
            'primary_trait_id': primary_trait_id,
            'primary_band':     primary_band,
            'trigger_trait_id': trigger_trait_id,
            'trigger_band':     trigger_band,
            'narrative':        narrative,
        })

    return interactions


def parse_endpoint_sheet(ws):
    """09_endpoints -> list of dict. One row per endpoint."""
    endpoints = []
    for row in list(ws.iter_rows(values_only=True))[2:]:   # row 1 = header, row 2 = description
        trait_id = _cell(row, 0)
        if not _is_valid_trait_id(trait_id):
            continue
        endpoints.append({
            'trait_id':       trait_id,
            'band':           _cell(row, 1),
            'endpoint_type':  _cell(row, 2),
            'endpoint_level': _cell(row, 3),
            'block_key':      _cell(row, 4),
            'note':           _cell(row, 5),
        })
    return endpoints


def parse_block_sheet(ws):
    """10_endpoint_blocks -> list of dict."""
    blocks = []
    for row in list(ws.iter_rows(values_only=True))[2:]:
        block_key = _cell(row, 0)
        if not block_key:
            continue
        blocks.append({
            'block_key':     block_key,
            'question_type': _cell(row, 1),
            'header_text':   _cell(row, 2),
            'sort_order':    int(_cell(row, 3)),
            'priority':      int(_cell(row, 4)),
            'footnote_rule': _cell(row, 5),
        })
    return blocks


def _find_sheet(wb, key):
    """Return the first sheet whose name contains key (case-insensitive)."""
    for name in wb.sheetnames:
        if key.lower() in name.lower():
            return name
    return None


def parse_excel(excel_path, require_endpoints=True):
    print(f"[Parse] Opening: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    sheets = wb.sheetnames
    print(f"[Parse] Sheets found: {sheets}")

    band_sheet_name = _find_sheet(wb, BAND_SHEET_KEY)
    interaction_sheet_name = _find_sheet(wb, INTERACTION_SHEET_KEY)

    if not band_sheet_name:
        raise ValueError(f"No sheet containing '{BAND_SHEET_KEY}' found. Available: {sheets}")
    if not interaction_sheet_name:
        raise ValueError(f"No sheet containing '{INTERACTION_SHEET_KEY}' found. Available: {sheets}")

    print(f"[Parse] Using band sheet: '{band_sheet_name}'")
    print(f"[Parse] Using interaction sheet: '{interaction_sheet_name}'")

    definitions, bands = parse_band_sheet(wb[band_sheet_name])
    interactions = parse_interaction_sheet(wb[interaction_sheet_name])

    # Endpoint sheets (V6.3+). Their absence must be loud: trait_endpoints has an
    # FK to trait_definitions, so the TRUNCATE ... CASCADE below would silently
    # empty it if we loaded an older spec file that has no 09 sheet.
    endpoint_sheet_name = _find_sheet(wb, ENDPOINT_SHEET_KEY)
    block_sheet_name = _find_sheet(wb, BLOCK_SHEET_KEY)
    if (not endpoint_sheet_name or not block_sheet_name) and not require_endpoints:
        print("[Parse] WARNING: no endpoint sheets; existing endpoint rows left untouched.")
        wb.close()
        print(f"[Parse] trait_definitions: {len(definitions)} records")
        print(f"[Parse] trait_bands:       {len(bands)} records")
        print(f"[Parse] trait_interactions:{len(interactions)} records")
        return definitions, bands, interactions, None, None
    if not endpoint_sheet_name or not block_sheet_name:
        raise ValueError(
            f"Spec file has no '{ENDPOINT_SHEET_KEY}' / '{BLOCK_SHEET_KEY}' sheet "
            f"(found: {sheets}). Use a V6.3+ spec, or pass --skip-endpoints to keep "
            f"the existing endpoint rows untouched. Refusing to continue silently: "
            f"a full migration would wipe trait_endpoints via TRUNCATE CASCADE.")

    endpoints = parse_endpoint_sheet(wb[endpoint_sheet_name])
    blocks = parse_block_sheet(wb[block_sheet_name])
    wb.close()

    print(f"[Parse] trait_definitions: {len(definitions)} records")
    print(f"[Parse] trait_bands:       {len(bands)} records")
    print(f"[Parse] trait_interactions:{len(interactions)} records")
    print(f"[Parse] trait_endpoints:   {len(endpoints)} records")
    print(f"[Parse] endpoint_blocks:   {len(blocks)} records")

    return definitions, bands, interactions, endpoints, blocks


# ---------------------------------------------------------------------------
# DB backup
# ---------------------------------------------------------------------------

def _escape_sql_value(v):
    if v is None:
        return 'NULL'
    if isinstance(v, (int, float)):
        return str(v)
    if isinstance(v, dict):
        s = json.dumps(v, ensure_ascii=False).replace("'", "''")
        return f"'{s}'"
    s = str(v).replace("'", "''")
    return f"'{s}'"


def backup_table(conn, table_name, f):
    from sqlalchemy import text as sa_text
    result = conn.execute(sa_text(f"SELECT * FROM {table_name}"))
    cols = list(result.keys())
    rows = result.mappings().fetchall()
    f.write(f"\n-- ========== {table_name} ({len(rows)} rows) ==========\n")
    f.write(f"DELETE FROM {table_name};\n")
    for row in rows:
        values = ', '.join(_escape_sql_value(row[c]) for c in cols)
        col_list = ', '.join(cols)
        f.write(f"INSERT INTO {table_name} ({col_list}) VALUES ({values});\n")
    print(f"[Backup] {table_name}: {len(rows)} rows written")


def create_backup(engine, backup_dir):
    os.makedirs(backup_dir, exist_ok=True)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    path = os.path.join(backup_dir, f'trait_backup_{ts}.sql')

    with engine.connect() as conn:
        with open(path, 'w', encoding='utf-8') as f:
            f.write(f"-- Trait tables backup generated {ts}\n")
            f.write("BEGIN;\n")
            backup_table(conn, 'trait_definitions', f)
            backup_table(conn, 'trait_bands', f)
            backup_table(conn, 'trait_interactions', f)
            f.write("\nCOMMIT;\n")

    print(f"[Backup] Written to: {path}")
    return path


# ---------------------------------------------------------------------------
# DB write
# ---------------------------------------------------------------------------

def apply_schema_migration(engine):
    """Add new columns if they don't exist yet (idempotent)."""
    from sqlalchemy import text as sa_text
    migrations = [
        "ALTER TABLE trait_definitions ADD COLUMN IF NOT EXISTS definition_en TEXT",
        "ALTER TABLE trait_definitions ADD COLUMN IF NOT EXISTS hidden_anchor TEXT",
        "ALTER TABLE trait_bands ADD COLUMN IF NOT EXISTS usage_note TEXT",
        "ALTER TABLE trait_bands ADD COLUMN IF NOT EXISTS trait_interaction_guide TEXT",
        "ALTER TABLE trait_bands ADD COLUMN IF NOT EXISTS version VARCHAR(20)",
    ]
    with engine.connect() as conn:
        for sql in migrations:
            conn.execute(sa_text(sql))
        conn.commit()
    print("[Schema] Column migrations applied.")


def apply_endpoint_schema(engine):
    """Create endpoint tables by executing the migration SQL file (idempotent)."""
    from sqlalchemy import text as sa_text
    if not os.path.exists(ENDPOINT_DDL_PATH):
        print(f"[Schema] WARNING: {ENDPOINT_DDL_PATH} not found; skipping endpoint DDL.")
        return
    body = open(ENDPOINT_DDL_PATH, encoding='utf-8').read().split('COMMIT;')[0].replace('BEGIN;', '')
    with engine.begin() as conn:
        for chunk in body.split(';'):
            stmt = '\n'.join(l for l in chunk.splitlines() if not l.strip().startswith('--')).strip()
            if stmt:
                conn.execute(sa_text(stmt))
    print("[Schema] Endpoint tables ensured (endpoint_blocks, trait_endpoints).")


def write_endpoints(engine, endpoints, blocks):
    """Reload endpoint_blocks (upsert) and trait_endpoints (replace)."""
    from sqlalchemy import text

    with engine.begin() as conn:
        for b in blocks:
            conn.execute(text("""
                INSERT INTO endpoint_blocks
                    (block_key, question_type, header_text, sort_order, priority, footnote_rule)
                VALUES
                    (:block_key, :question_type, :header_text, :sort_order, :priority, :footnote_rule)
                ON CONFLICT (block_key) DO UPDATE SET
                    question_type = EXCLUDED.question_type,
                    header_text   = EXCLUDED.header_text,
                    sort_order    = EXCLUDED.sort_order,
                    priority      = EXCLUDED.priority,
                    footnote_rule = EXCLUDED.footnote_rule
            """), b)
        print(f"[Write] endpoint_blocks: {len(blocks)} rows upserted")

        conn.execute(text("DELETE FROM trait_endpoints"))
        for e in endpoints:
            conn.execute(text("""
                INSERT INTO trait_endpoints
                    (trait_id, band, endpoint_type, endpoint_level, block_key, note)
                VALUES
                    (:trait_id, :band, :endpoint_type, :endpoint_level, :block_key, :note)
            """), e)
        print(f"[Write] trait_endpoints: {len(endpoints)} rows inserted")


def verify_endpoints(engine, endpoints):
    """Post-load consistency check. Loud failure beats a silently shrunken set."""
    from sqlalchemy import text
    from collections import Counter
    with engine.connect() as conn:
        n = conn.execute(text("SELECT count(*) FROM trait_endpoints")).scalar()
        orphan = conn.execute(text("""
            SELECT count(*) FROM trait_endpoints e
            LEFT JOIN endpoint_blocks b USING (block_key) WHERE b.block_key IS NULL
        """)).scalar()
        by_type = conn.execute(text(
            "SELECT endpoint_type, count(*) FROM trait_endpoints GROUP BY 1 ORDER BY 1")).fetchall()
    problems = []
    if n != len(endpoints):
        problems.append(f"row count {n} != spec {len(endpoints)}")
    if orphan:
        problems.append(f"{orphan} rows reference an unknown block_key")
    expected = Counter(e['endpoint_type'] for e in endpoints)
    for t, c in by_type:
        if expected.get(t) != c:
            problems.append(f"type '{t}': db {c} != spec {expected.get(t)}")
    print(f"[Verify] trait_endpoints: {n} rows, by type {dict(by_type)}, orphan block_key {orphan}")
    if problems:
        raise RuntimeError("Endpoint verification failed: " + "; ".join(problems))
    print("[Verify] Endpoint data consistent with spec.")


def write_to_db(engine, definitions, bands, interactions):
    from sqlalchemy import text

    with engine.begin() as conn:
        # Truncate in FK-dependency order
        conn.execute(text("TRUNCATE trait_interactions, trait_bands, trait_definitions CASCADE"))
        print("[Write] Tables truncated.")

        # Insert trait_definitions
        for d in definitions:
            conn.execute(text("""
                INSERT INTO trait_definitions
                    (trait_id, name_zh, name_en, dimension, definition, definition_en, hidden_anchor)
                VALUES
                    (:trait_id, :name_zh, :name_en, :dimension, :definition, :definition_en, :hidden_anchor)
            """), d)
        print(f"[Write] trait_definitions: {len(definitions)} rows inserted")

        # Insert trait_bands
        for b in bands:
            b_copy = dict(b)
            b_copy['ai_guidance'] = json.dumps(b_copy['ai_guidance'], ensure_ascii=False)
            conn.execute(text("""
                INSERT INTO trait_bands
                    (trait_id, band, min_score, max_score, semantic_label, description,
                     management_focus, usage_note, trait_interaction_guide,
                     report_wording, report_wording_friendly, trait_project, ai_guidance, version)
                VALUES
                    (:trait_id, :band, :min_score, :max_score, :semantic_label, :description,
                     :management_focus, :usage_note, :trait_interaction_guide,
                     :report_wording, :report_wording_friendly, :trait_project,
                     CAST(:ai_guidance AS jsonb), :version)
            """), b_copy)
        print(f"[Write] trait_bands: {len(bands)} rows inserted")

        # Insert trait_interactions — only where primary_trait_id exists in definitions
        valid_ids = {d['trait_id'] for d in definitions}
        filtered = [i for i in interactions if i['primary_trait_id'] in valid_ids]
        skipped = len(interactions) - len(filtered)
        if skipped:
            print(f"[Write] Skipping {skipped} interactions with unknown primary_trait_id")
        for i in filtered:
            conn.execute(text("""
                INSERT INTO trait_interactions
                    (primary_trait_id, primary_band, trigger_trait_id, trigger_band, narrative)
                VALUES
                    (:primary_trait_id, :primary_band, :trigger_trait_id, :trigger_band, :narrative)
            """), i)
        print(f"[Write] trait_interactions: {len(filtered)} rows inserted")

    print("[Write] Commit successful.")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description='Migrate trait data from Excel to PostgreSQL')
    parser.add_argument('--excel', required=True, help='Path to Traitty_RAG_SpeC Excel file')
    parser.add_argument('--dry-run', action='store_true', help='Parse only, do not write to DB')
    parser.add_argument('--endpoints-only', action='store_true',
                        help='Load only 09_endpoints / 10_endpoint_blocks; leave the three trait tables untouched')
    parser.add_argument('--skip-endpoints', action='store_true',
                        help='Allow a pre-V6.3 spec with no endpoint sheets; existing endpoint rows are left as-is')
    args = parser.parse_args()

    excel_path = os.path.abspath(args.excel)
    if not os.path.exists(excel_path):
        print(f"[ERROR] File not found: {excel_path}")
        sys.exit(1)

    definitions, bands, interactions, endpoints, blocks = parse_excel(
        excel_path, require_endpoints=not args.skip_endpoints)

    if args.dry_run:
        print("\n[DRY RUN] Parsed counts:")
        print(f"  trait_definitions : {len(definitions)}")
        print(f"  trait_bands       : {len(bands)}")
        print(f"  trait_interactions: {len(interactions)}")
        print(f"  trait_endpoints   : {len(endpoints) if endpoints is not None else 'n/a'}")
        print(f"  endpoint_blocks   : {len(blocks) if blocks is not None else 'n/a'}")
        if definitions:
            print(f"\nSample definition: {definitions[0]}")
        if bands:
            b = dict(bands[0])
            b['ai_guidance'] = str(b['ai_guidance'])[:80]
            print(f"Sample band      : {b}")
        if interactions:
            print(f"Sample interaction: {interactions[0]}")
        print("\n[DRY RUN] Done. No data written.")
        return

    engine = get_db_engine()

    if args.endpoints_only:
        if endpoints is None:
            print("[ERROR] --endpoints-only needs a spec file with endpoint sheets.")
            sys.exit(1)
        print("\n[Step 1] Ensuring endpoint schema...")
        apply_endpoint_schema(engine)
        print("\n[Step 2] Writing endpoint data (trait tables untouched)...")
        write_endpoints(engine, endpoints, blocks)
        print("\n[Step 3] Verifying...")
        verify_endpoints(engine, endpoints)
        print("\n[DONE] Endpoint tables loaded.")
        return

    backup_dir = os.path.join(os.path.dirname(__file__), 'backups')
    print("\n[Step 1] Creating SQL backup...")
    backup_path = create_backup(engine, backup_dir)

    print("\n[Step 2] Applying schema migration...")
    apply_schema_migration(engine)
    apply_endpoint_schema(engine)

    print("\n[Step 3] Writing new data...")
    write_to_db(engine, definitions, bands, interactions)

    # After write_to_db: TRUNCATE ... CASCADE has emptied trait_endpoints, so it
    # must be reloaded in the same run.
    if endpoints is not None:
        print("\n[Step 4] Writing endpoint data...")
        write_endpoints(engine, endpoints, blocks)
        verify_endpoints(engine, endpoints)

    print(f"\n[DONE] Migration complete. Backup saved at: {backup_path}")


if __name__ == '__main__':
    main()
