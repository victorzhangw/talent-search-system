"""Monthly usage workbook for one environment (default PRD).

Usage:
    python scripts/generate_monthly_usage_report.py                 # PRD -> docs/customer/
    python scripts/generate_monthly_usage_report.py --db uat
    python scripts/generate_monthly_usage_report.py --out some.xlsx

Sheets
------
1. 各月使用次數 -- one row per calendar month. Column B (使用次數) is the headline
   metric and counts every row in chat_sessions by started_at, including sessions
   that were opened and abandoned without a message. The remaining columns are
   reference figures so the headline number can be sanity-checked.
2. 長條圖 -- a native Excel bar chart whose series points at Sheet1!B, so editing a
   number on sheet 1 redraws the bar. Nothing is baked into the picture.
3. 資料來源說明 -- provenance, and the caveat about extending the chart range.

Times are Taiwan local (UTC+8). chat_sessions.started_at is stored naive UTC, the same
convention generate_token_usage_report.py uses.

Read-only: the connection is opened readonly, so this cannot write to the database
no matter what the SQL says.
"""

import argparse
import os
import sys
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPT_DIR)

try:
    import openpyxl
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter
except ImportError:
    print('[ERROR] openpyxl not installed. Run: pip install openpyxl')
    sys.exit(1)

from purge_user_chat_history import DBS, connect, load_env

REPO_ROOT = os.path.join(SCRIPT_DIR, '..', '..')

# (header, key, width). The chart plots the second column; the rest are reference.
COLUMNS = [
    ('月份', 'month', 12),
    ('使用次數 (對話次數)', 'sessions', 20),
    ('其中有實際對話', 'sessions_with_msg', 16),
    ('提問則數', 'user_msgs', 12),
    ('回覆則數', 'assistant_msgs', 12),
    ('Token 數', 'tokens', 14),
    ('使用帳號數', 'accounts', 12),
]


def q(cur, sql, args=None):
    cur.execute(sql, args) if args else cur.execute(sql)
    return cur.fetchall()


def month_range(first, last):
    """Every YYYY-MM from first to last inclusive, so a month with no use still
    gets a bar at zero instead of silently collapsing the time axis."""
    y, m = (int(x) for x in first.split('-'))
    ly, lm = (int(x) for x in last.split('-'))
    out = []
    while (y, m) <= (ly, lm):
        out.append(f'{y:04d}-{m:02d}')
        y, m = (y + 1, 1) if m == 12 else (y, m + 1)
    return out


def fetch(cur):
    """Per-month figures, keyed on the session's own start month.

    Message counts are attributed to the month the *session* started, not the month
    the message was sent, so every column on a row describes the same set of
    sessions. A session that spans midnight on the 1st is rare here (the longest
    runs minutes) but this keeps the row internally consistent regardless.
    """
    rows = q(cur, """
        SELECT to_char(s.started_at + interval '8 hours', 'YYYY-MM')      AS ym,
               count(*)                                                    AS sessions,
               count(*) FILTER (WHERE m.total > 0)                         AS with_msg,
               coalesce(sum(m.user_msgs), 0)                               AS user_msgs,
               coalesce(sum(m.assistant_msgs), 0)                          AS assistant_msgs,
               coalesce(sum(m.tokens), 0)                                  AS tokens,
               count(DISTINCT s.user_id)                                   AS accounts
        FROM chat_sessions s
        LEFT JOIN LATERAL (
            SELECT count(*)                                       AS total,
                   count(*) FILTER (WHERE role = 'user')          AS user_msgs,
                   count(*) FILTER (WHERE role = 'assistant')     AS assistant_msgs,
                   coalesce(sum(token_usage), 0)                  AS tokens
            FROM chat_messages WHERE session_id = s.session_id
        ) m ON true
        GROUP BY ym ORDER BY ym""")
    by_month = {r[0]: dict(zip([c[1] for c in COLUMNS[1:]], r[1:])) for r in rows}
    if not by_month:
        return []
    months = month_range(min(by_month), max(by_month))
    blank = {c[1]: 0 for c in COLUMNS[1:]}
    return [{'month': ym, **by_month.get(ym, blank)} for ym in months]


def build_workbook(data, db_label, dbname):
    wb = openpyxl.Workbook()
    bold = Font(bold=True)
    n = len(data)

    ws1 = wb.active
    ws1.title = '各月使用次數'
    ws1.append([c[0] for c in COLUMNS])
    for i, (header, _key, width) in enumerate(COLUMNS, start=1):
        ws1.cell(row=1, column=i).font = bold
        ws1.cell(row=1, column=i).alignment = Alignment(horizontal='center')
        ws1.column_dimensions[get_column_letter(i)].width = width
    for row in data:
        ws1.append([row[c[1]] for c in COLUMNS])
    ws1.freeze_panes = 'A2'

    # One blank row, so the total sits outside the chart's data range and editing it
    # cannot move a bar.
    total_row = n + 3
    ws1.cell(row=total_row, column=1, value='合計').font = bold
    for i in range(2, len(COLUMNS) + 1):
        col = get_column_letter(i)
        # 使用帳號數 is a distinct count; summing it down the column would double-count
        # an account that appears in several months.
        if COLUMNS[i - 1][1] == 'accounts':
            ws1.cell(row=total_row, column=i, value='(不適用，同一帳號跨月會重複計算)')
        else:
            ws1.cell(row=total_row, column=i, value=f'=SUM({col}2:{col}{n + 1})').font = bold

    ws2 = wb.create_sheet('長條圖')
    chart = BarChart()
    chart.type = 'col'
    chart.title = f'{db_label} 各月使用次數（對話次數）'
    chart.y_axis.title = '使用次數'
    chart.x_axis.title = '月份'
    chart.height = 11
    chart.width = 24
    chart.gapWidth = 60
    # Absolute references into sheet 1: change B2:B7 there and the bars follow.
    chart.add_data(Reference(ws1, min_col=2, min_row=1, max_row=n + 1), titles_from_data=True)
    chart.set_categories(Reference(ws1, min_col=1, min_row=2, max_row=n + 1))
    ws2.add_chart(chart, 'B2')
    ws2['A1'] = '本圖的資料來源是「各月使用次數」工作表的 B 欄，修改該欄數字後圖形會自動更新。'
    ws2['A1'].font = Font(italic=True)

    ws3 = wb.create_sheet('資料來源說明')
    notes = [
        ('產出時間', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        ('環境', f'{db_label} ({dbname})'),
        ('資料來源', 'PostgreSQL chat_sessions / chat_messages'),
        ('統計期間', f'{data[0]["month"]} ~ {data[-1]["month"]}（依 chat_sessions.started_at）'),
        ('時區', 'Taiwan 本地時間 UTC+8；資料庫存的是 naive UTC，查詢時加 8 小時'),
        ('', ''),
        ('「使用次數」定義',
         'chat_sessions 的資料列數，即使用者開啟對話的次數。含開啟後未送出任何訊息就離開的 session；'
         'C 欄「其中有實際對話」為扣掉這類空 session 後的數字，兩者的差距即為空開次數。'),
        ('提問／回覆則數', 'chat_messages 中 role=user 與 role=assistant 的則數，歸屬到該 session 起始的月份。'),
        ('Token 數', 'chat_messages.token_usage 加總。僅供用量趨勢參考，計費依據仍以供應商官方用量報表為準。'),
        ('使用帳號數', '該月出現的相異 chat_sessions.user_id 個數。跨月會重複，故不做總計。'),
        ('', ''),
        ('修改數字',
         '長條圖的資料範圍是絕對參照 B2:B' + str(n + 1) + '。直接改這個範圍內的數字，圖會跟著變；'
         '但若要「新增月份列」，需一併把圖表的資料範圍往下拉，否則新列不會進圖。'),
    ]
    for k, v in notes:
        ws3.append([k, v])
    for row in ws3.iter_rows(min_col=1, max_col=1):
        row[0].font = bold
    ws3.column_dimensions['A'].width = 18
    ws3.column_dimensions['B'].width = 100
    for row in ws3.iter_rows(min_col=2, max_col=2):
        row[0].alignment = Alignment(wrap_text=True, vertical='top')

    return wb


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--db', choices=sorted(DBS), default='prd')
    ap.add_argument('--out', default=None)
    args = ap.parse_args()

    dbname = DBS[args.db]
    label = args.db.upper()
    conn = connect(load_env(), dbname, readonly=True)
    try:
        data = fetch(conn.cursor())
    finally:
        conn.close()

    if not data:
        print('[WARN] No sessions found; nothing written.')
        return 1

    print(f'{label} {dbname}')
    head = f'  {"月份":<9}' + ''.join(f'{c[0]:>14}' for c in COLUMNS[1:])
    print(head)
    for row in data:
        print(f'  {row["month"]:<11}' + ''.join(f'{row[c[1]]:>14,}' for c in COLUMNS[1:]))
    print(f'  {"合計":<10}' + ''.join(
        f'{sum(r[c[1]] for r in data):>14,}' if c[1] != 'accounts' else f'{"-":>14}'
        for c in COLUMNS[1:]))

    out = args.out or os.path.join(
        REPO_ROOT, 'docs', 'customer', f'{label}-月用量統計-{datetime.now():%Y%m%d}.xlsx')
    os.makedirs(os.path.dirname(os.path.abspath(out)), exist_ok=True)
    build_workbook(data, label, dbname).save(out)
    print(f'\n[OK] Wrote {os.path.abspath(out)}')
    return 0


if __name__ == '__main__':
    sys.exit(main())
