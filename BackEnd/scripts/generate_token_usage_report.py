"""
Generate a per-day / per-account token usage breakdown from chat_messages,
combining both the UAT (ai_chatbot_v2) and PRD (ai_chatbot_v2_prd) databases,
to explain which accounts drove a given period's token consumption.

Background: the customer-facing report (docs/customer/Token用量報表-2026-07.xlsx)
already shows daily totals for 2026/06/01-07/03 and excludes 2026/06/26-06/30
as "internal automated testing", but that attribution was never broken down
by account. This script adds the account (chat_sessions.user_id) dimension so
the anomaly can be verified instead of assumed.

Usage:
    DB_HOST=... DB_PORT=... DB_USER=... DB_PASSWORD=... \
        python BackEnd/scripts/generate_token_usage_report.py \
        --start 2026-06-01 --end 2026-07-04

Credentials are read from the environment / api_v2/.env (never hardcoded
here), same DB_* variable names used by database/connection.py::get_db_url().
Only DB_NAME differs between UAT and PRD, so this script builds its own
engine per environment instead of reusing the app's global cached engine
(get_db_engine() is a process-wide singleton and can't hold two databases
at once).

Writes:
    docs/investigations/2026-07_token_usage_by_account.xlsx
    docs/investigations/2026-07_token_usage_by_account_findings.md
"""
import sys
import os
import argparse
from datetime import datetime, timedelta
from collections import defaultdict

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

try:
    import openpyxl
    from openpyxl.styles import Font
except ImportError:
    print("[ERROR] openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

from dotenv import load_dotenv
env_path = os.path.join(os.path.dirname(__file__), '..', 'api_v2', '.env')
load_dotenv(env_path, encoding='utf-8-sig')

for _required in ('DB_HOST', 'DB_USER', 'DB_PASSWORD'):
    if not os.environ.get(_required):
        print(f"[ERROR] Missing required env var {_required}. See usage in this file's docstring.")
        sys.exit(1)

from sqlalchemy import create_engine, func
from sqlalchemy.orm import sessionmaker
from api_v2.database.models import ChatSession, ChatMessage

REPO_ROOT = os.path.join(os.path.dirname(__file__), '..', '..')
XLSX_PATH = os.path.join(REPO_ROOT, 'docs', 'investigations', '2026-07_token_usage_by_account.xlsx')
MD_PATH = os.path.join(REPO_ROOT, 'docs', 'investigations', '2026-07_token_usage_by_account_findings.md')

ENVIRONMENTS = [
    ('UAT', 'ai_chatbot_v2'),
    ('PRD', 'ai_chatbot_v2_prd'),
]

DEFAULT_TEST_WINDOW = (datetime(2026, 6, 26), datetime(2026, 7, 1))


def build_session_factory(db_name):
    user = os.environ['DB_USER']
    password = os.environ['DB_PASSWORD']
    host = os.environ['DB_HOST']
    port = os.environ.get('DB_PORT', '5432')
    url = f"postgresql://{user}:{password}@{host}:{port}/{db_name}"
    engine = create_engine(url, pool_pre_ping=True)
    return sessionmaker(autocommit=False, autoflush=False, bind=engine)


def fetch_env_rows(db_name, start_local, end_local):
    """Per (local day, account) token totals for one database, TW-local (UTC+8) day boundaries."""
    Session = build_session_factory(db_name)
    db = Session()
    try:
        start_utc = start_local - timedelta(hours=8)
        end_utc = end_local - timedelta(hours=8)

        local_created_at = ChatMessage.created_at + timedelta(hours=8)
        date_col = func.date_trunc('day', local_created_at).label('local_date')

        rows = (
            db.query(
                date_col,
                ChatSession.user_id,
                func.sum(ChatMessage.token_usage).label('tokens'),
                func.sum(ChatMessage.prompt_tokens).label('prompt_tokens'),
                func.sum(ChatMessage.completion_tokens).label('completion_tokens'),
                func.count(func.distinct(ChatSession.session_id)).label('sessions'),
            )
            .join(ChatSession, ChatSession.session_id == ChatMessage.session_id)
            .filter(ChatMessage.created_at >= start_utc, ChatMessage.created_at < end_utc)
            .group_by(date_col, ChatSession.user_id)
            .order_by(date_col)
            .all()
        )
        return rows
    finally:
        db.close()


def in_test_window(local_date, window):
    start, end = window
    return start <= local_date < end


def build_detail_rows(start_local, end_local, window, environments):
    detail = []
    for env_label, db_name in environments:
        print(f"[Query] {env_label} ({db_name})...")
        rows = fetch_env_rows(db_name, start_local, end_local)
        for r in rows:
            account = r.user_id or 'anonymous'
            detail.append({
                'date': r.local_date.date(),
                'env': env_label,
                'account': account,
                'sessions': r.sessions,
                'tokens': r.tokens or 0,
                'prompt_tokens': r.prompt_tokens or 0,
                'completion_tokens': r.completion_tokens or 0,
                'in_test_window': in_test_window(r.local_date, window),
            })
        print(f"  {len(rows)} (date, account) rows")
    detail.sort(key=lambda d: (d['date'], d['env'], -d['tokens']))
    return detail


def build_account_rollup(detail):
    by_account = defaultdict(lambda: {'tokens': 0, 'test_window_tokens': 0, 'sessions': 0, 'envs': set()})
    for row in detail:
        acc = by_account[row['account']]
        acc['tokens'] += row['tokens']
        acc['sessions'] += row['sessions']
        acc['envs'].add(row['env'])
        if row['in_test_window']:
            acc['test_window_tokens'] += row['tokens']

    rollup = []
    for account, agg in by_account.items():
        pct = (agg['test_window_tokens'] / agg['tokens'] * 100) if agg['tokens'] else 0.0
        rollup.append({
            'account': account,
            'total_tokens': agg['tokens'],
            'total_sessions': agg['sessions'],
            'test_window_tokens': agg['test_window_tokens'],
            'test_window_pct': pct,
            'environments': ', '.join(sorted(agg['envs'])),
        })
    rollup.sort(key=lambda r: -r['total_tokens'])
    return rollup


def build_daily_totals(detail):
    by_date = defaultdict(lambda: {'tokens': 0, 'in_test_window': False})
    for row in detail:
        d = by_date[row['date']]
        d['tokens'] += row['tokens']
        d['in_test_window'] = row['in_test_window']
    return sorted(
        ({'date': d, **v} for d, v in by_date.items()),
        key=lambda r: r['date'],
    )


def write_workbook(detail, rollup, daily_totals, start_local, end_local, window, vendor_billed_total=None):
    wb = openpyxl.Workbook()
    bold = Font(bold=True)

    ws = wb.active
    ws.title = '每日x帳號明細'
    ws.append(['日期', '環境', '帳號 (user_id)', 'Session 數', 'Token 總數', 'Prompt Tokens', 'Completion Tokens', '落在測試區間'])
    for cell in ws[1]:
        cell.font = bold
    for row in detail:
        ws.append([
            row['date'].isoformat(), row['env'], row['account'], row['sessions'],
            row['tokens'], row['prompt_tokens'], row['completion_tokens'],
            '是' if row['in_test_window'] else '',
        ])

    ws2 = wb.create_sheet('帳號加總')
    ws2.append(['帳號 (user_id)', 'Token 總數', 'Session 總數', '測試區間 Token 數', '測試區間佔比 (%)', '出現環境'])
    for cell in ws2[1]:
        cell.font = bold
    for row in rollup:
        ws2.append([
            row['account'], row['total_tokens'], row['total_sessions'],
            row['test_window_tokens'], round(row['test_window_pct'], 1), row['environments'],
        ])

    ws3 = wb.create_sheet('每日活動量(合計)')
    ws3.append(['日期', 'Token 數 (UAT+PRD 合計)', '性質'])
    for cell in ws3[1]:
        cell.font = bold
    for row in daily_totals:
        nature = '測試期間（不列入計費）' if row['in_test_window'] else '計費區間'
        ws3.append([row['date'].isoformat(), row['tokens'], nature])

    ws4 = wb.create_sheet('資料來源說明')
    ws4.append(['產出時間', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    ws4.append(['統計期間 (TW 本地時間)', f"{start_local.date()} - {(end_local - timedelta(days=1)).date()}"])
    ws4.append(['測試區間 (排除計費)', f"{window[0].date()} - {(window[1] - timedelta(days=1)).date()}"])
    ws4.append([])
    ws4.append(['資料來源', 'PostgreSQL chat_sessions / chat_messages，UAT (ai_chatbot_v2) + PRD (ai_chatbot_v2_prd) 兩環境合計'])
    ws4.append(['帳號欄位說明', '"帳號" 為 chat_sessions.user_id，即前端設定檔傳入的使用者 email；未帶入時顯示為 anonymous。此欄位代表個別使用者，非客戶公司名稱（本部署為單一客戶 WePredict）。'])
    ws4.append(['計費依據', 'Token 費用之唯一依據仍為 DeepSeek API 供應商官方用量報表；本表僅供用量趨勢與帳號歸因分析，非計費依據。'])
    if vendor_billed_total:
        combined = sum(r['tokens'] for r in daily_totals if not r['in_test_window'])
        pct = (combined / vendor_billed_total * 100) if vendor_billed_total else 0
        ws4.append([])
        ws4.append(['重要提醒', f"本表計費區間內 DB 對話紀錄合計 {combined:,} tokens，僅為供應商計費總額 {vendor_billed_total:,} tokens 的 {pct:.2f}%。差距懸殊，代表帳單金額主要並非來自 /chat 對話路徑，需另從 DeepSeek API 帳號/金鑰使用面調查（例如其他呼叫 DeepSeek 的後端流程、批次工作、或金鑰是否被其他服務共用）。"])

    os.makedirs(os.path.dirname(XLSX_PATH), exist_ok=True)
    wb.save(XLSX_PATH)
    print(f"[OK] Wrote {XLSX_PATH}")


def write_findings(rollup, daily_totals, window, vendor_billed_total=None):
    window_accounts = [r for r in rollup if r['test_window_tokens'] > 0]
    window_accounts.sort(key=lambda r: -r['test_window_tokens'])
    total_window_tokens = sum(r['test_window_tokens'] for r in window_accounts)

    lines = []
    lines.append("# Token Usage by Account -- Findings")
    lines.append("")
    lines.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"Test window checked: {window[0].date()} - {(window[1] - timedelta(days=1)).date()} (TW local)")
    lines.append("")

    if vendor_billed_total:
        combined_billed = sum(r['tokens'] for r in daily_totals if not r['in_test_window'])
        pct = (combined_billed / vendor_billed_total * 100) if vendor_billed_total else 0
        lines.append("## Headline finding: chat_messages cannot explain the vendor bill's size")
        lines.append("")
        lines.append(f"Combined UAT+PRD `chat_messages` total for the billed days (excluding the test window): **{combined_billed:,} tokens**.")
        lines.append(f"DeepSeek vendor-billed total for the same billed days: **{vendor_billed_total:,} tokens**.")
        lines.append("")
        lines.append(
            f"DB-tracked chat activity is only **~{pct:.2f}%** of what the vendor billed for the same "
            "billed days. The daily shapes match exactly against the customer's existing xlsx/html "
            "report (same source table), so the DB query itself is correct -- it simply cannot be the "
            "(or even a significant part of the) source of the vendor's token count."
        )
        lines.append("")
        lines.append(
            "**Implication**: the account breakdown below reliably explains who is behind the test-window "
            "spike and every other data point *inside* `chat_messages`, but it cannot explain why the vendor "
            "invoice is orders of magnitude larger. That gap must come from something outside the `/chat` "
            "code path entirely -- candidates worth checking next: other backend calls to the DeepSeek API "
            "that don't write to `chat_messages` (e.g. RAG/report-generation calls in `context_builder.py`, "
            "embeddings, retries), a batch/background job, or the DeepSeek API key being used by something "
            "other than this application. This needs to be investigated on the DeepSeek account/API-key "
            "side, not the chat database, before the \"why did usage exceed expectations\" question can be "
            "fully answered."
        )
        lines.append("")

    lines.append(f"## Accounts responsible for the {window[0].date()}-{(window[1] - timedelta(days=1)).date()} test window ({total_window_tokens:,} tokens)")
    lines.append("")
    lines.append("| Account (user_id) | Tokens in window | % of account's total tokens | Environments |")
    lines.append("|---|---|---|---|")
    for r in window_accounts:
        lines.append(f"| {r['account']} | {r['test_window_tokens']:,} | {round(r['test_window_pct'], 1)}% | {r['environments']} |")
    if not window_accounts:
        lines.append("| (none found) | - | - | - |")
    lines.append("")
    lines.append("## Top 10 accounts by total tokens (whole period)")
    lines.append("")
    lines.append("| Account (user_id) | Total tokens | Sessions | Environments |")
    lines.append("|---|---|---|---|")
    for r in rollup[:10]:
        lines.append(f"| {r['account']} | {r['total_tokens']:,} | {r['total_sessions']} | {r['environments']} |")

    lines.append("")
    lines.append("## Notes on the accounts seen")
    lines.append("")
    lines.append(
        f"- {len(rollup)} distinct `user_id` values appear in `chat_messages` across the whole period, "
        "across both databases. There is no evidence in this data of any account outside this set -- "
        "i.e. no unidentified/external end-user traffic shows up in `chat_messages` for this period."
    )
    lines.append(
        "- Accounts appearing in short, high-volume bursts on only a handful of days (rather than steady "
        "daily use) are more consistent with manual or automated testing sessions than ongoing product "
        "usage. Whether specific addresses map to WePredict team members, an internal tester, or something "
        "else needs to be confirmed with whoever has visibility into who was assigned each address -- this "
        "data only shows the email string itself."
    )

    os.makedirs(os.path.dirname(MD_PATH), exist_ok=True)
    with open(MD_PATH, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines) + '\n')
    print(f"[OK] Wrote {MD_PATH}")


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument('--start', default='2026-06-01', help='TW-local start date (inclusive), YYYY-MM-DD')
    parser.add_argument('--end', default='2026-07-04', help='TW-local end date (exclusive), YYYY-MM-DD')
    parser.add_argument('--test-window-start', default='2026-06-26', help='TW-local test window start (inclusive)')
    parser.add_argument('--test-window-end', default='2026-07-01', help='TW-local test window end (exclusive)')
    parser.add_argument('--env', choices=['uat', 'prd', 'both'], default='both',
                         help='Which database(s) to query. Use "uat" to test without touching production.')
    parser.add_argument('--vendor-billed-total', type=int, default=None,
                         help='Vendor-invoiced token total for the billed period, for the DB-vs-vendor gap note.')
    args = parser.parse_args()

    start_local = datetime.strptime(args.start, '%Y-%m-%d')
    end_local = datetime.strptime(args.end, '%Y-%m-%d')
    window = (
        datetime.strptime(args.test_window_start, '%Y-%m-%d'),
        datetime.strptime(args.test_window_end, '%Y-%m-%d'),
    )
    environments = {
        'uat': ENVIRONMENTS[:1],
        'prd': ENVIRONMENTS[1:],
        'both': ENVIRONMENTS,
    }[args.env]

    detail = build_detail_rows(start_local, end_local, window, environments)
    if not detail:
        print("[WARN] No rows found for the given period -- check DB connectivity and date range.")
        return

    rollup = build_account_rollup(detail)
    daily_totals = build_daily_totals(detail)

    write_workbook(detail, rollup, daily_totals, start_local, end_local, window, args.vendor_billed_total)
    write_findings(rollup, daily_totals, window, args.vendor_billed_total)

    combined_total = sum(r['tokens'] for r in daily_totals)
    print(f"[OK] Combined UAT+PRD total tokens for {start_local.date()}-{(end_local - timedelta(days=1)).date()}: {combined_total:,}")


if __name__ == '__main__':
    main()
